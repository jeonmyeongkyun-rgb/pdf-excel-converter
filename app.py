import ssl
ssl._create_default_https_context = ssl._create_unverified_context

import streamlit as st
import google.generativeai as genai
import fitz  # PyMuPDF
import pandas as pd
import os
import io
import zipfile
import glob
import re
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# ==============================================================================
# [필수] 구글 AI Studio에서 발급받은 키를 여기에 넣으세요
GOOGLE_API_KEY = "AIzaSyDAGuC0v4hhdwegQhlxNWwAPwe3Vaym0rQ"
# ==============================================================================

genai.configure(api_key=GOOGLE_API_KEY)

# 모델 설정 (2.5 우선, 없으면 1.5)
try:
    model = genai.GenerativeModel('gemini-2.5-flash')
except:
    try:
        model = genai.GenerativeModel('gemini-1.5-flash')
    except:
        st.error("❌ 모델 로딩 실패.")

st.set_page_config(page_title="Premium PDF Converter", page_icon="💎", layout="wide")

# --- 파일 청소 ---
def clean_up_trash():
    for f in glob.glob("temp_*.pdf") + glob.glob("*.xlsx"):
        try: os.remove(f)
        except: pass
clean_up_trash()

# --- 디자인 ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;700&display=swap');
    .stApp { background-color: #F1F5F9; font-family: 'Noto Sans KR', sans-serif; color: #334155; }
    h1 { color: #0F172A; font-weight: 800; text-align: center; letter-spacing: -1px; }
    .subtitle { text-align: center; color: #64748B; margin-bottom: 2rem; }
    [data-testid='stFileUploader'] { background: white; border: 2px dashed #94A3B8; border-radius: 16px; padding: 40px; }
    div.stButton > button { background-color: #2563EB; color: white; border: none; border-radius: 8px; padding: 0.8rem; width: 100%; font-weight: bold; font-size: 1rem; box-shadow: 0 4px 6px rgba(37, 99, 235, 0.2); }
    div.stButton > button:hover { background-color: #1D4ED8; transform: translateY(-2px); }
    .stSuccess, .stError { border-radius: 8px; font-weight: 500; }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1>Premium PDF ➡️ Excel</h1>", unsafe_allow_html=True)
st.markdown("<div class='subtitle'>상단 정보 포함 + 칸 밀림 방지 + 금액 콤마 유지</div>", unsafe_allow_html=True)

if "여기에" in GOOGLE_API_KEY:
    st.error("🚨 코드 16번째 줄에 API 키를 입력해주세요!")
    st.stop()

if 'processed_files' not in st.session_state:
    st.session_state.processed_files = []
if 'last_uploaded_ids' not in st.session_state:
    st.session_state.last_uploaded_ids = ""

uploaded_files = st.file_uploader("PDF 파일을 여기에 드래그하세요", type="pdf", accept_multiple_files=True)

def process_pdf_smart(file_bytes, original_name):
    temp_input = f"temp_{original_name}"
    file_root = os.path.splitext(original_name)[0]
    final_output_xls = f"{file_root}.xlsx"
    
    with open(temp_input, "wb") as f:
        f.write(file_bytes)

    try:
        doc = fitz.open(temp_input)
        
        # 헤더 정보와 테이블 데이터를 담을 리스트
        all_header_lines = [] 
        all_table_rows = []
        
        # 엑셀 컬럼 정의 (총 10개)
        columns = ["거래일자", "거래시간", "상태", "거래구분", "거래금액", "잔액", "취급점", "적요", "은행명", "상대계좌"]

        for i, page in enumerate(doc):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_data = pix.tobytes("png")
            image_parts = [{"mime_type": "image/png", "data": img_data}]
            
            # === [최종 수정 프롬프트] ===
            # 칸 밀림을 막기 위해 파이프(|) 개수를 강제하고, 헤더 정보도 가져오도록 지시
            prompt = """
            이 이미지 전체를 분석해서 데이터를 추출해줘.
            
            [구역 1: 상단 헤더 정보]
            - 문서 맨 위에 있는 '예금주', '계좌번호', '조회기간', '상품명' 등의 정보를 찾아서 그대로 텍스트로 적어줘.
            - 각 정보 앞에는 'HEADER:' 라는 태그를 붙여줘. (예: HEADER: 예금주 홍길동)
            
            [구역 2: 거래내역 표]
            - 표 안의 데이터를 파이프(|)로 구분해서 출력해.
            - [중요] 각 줄은 반드시 10개의 칸을 가져야 해. 빈 칸은 비워둬 (파이프 사이 공백).
            - 순서: 날짜|시간|상태|구분|거래금액|잔액|취급점|적요|은행명|상대계좌
            
            [엄격한 규칙]
            1. 금액과 잔액의 쉼표(,)는 **절대 지우지 마**. 그대로 유지해. (예: 10,000)
            2. 표면잔액과 취급점이 붙어있으면 반드시 파이프(|)로 나눠.
            3. 계좌번호는 숫자만 남겨.
            4. 배경의 'KB' 로고는 무시해.
            5. '거래일자', '거래시간' 같은 표의 제목 줄은 출력하지 마.
            """
            
            response = model.generate_content([prompt, image_parts[0]])
            
            if not response.text: continue

            # 줄 단위로 분석
            lines = response.text.strip().split('\n')
            
            for line in lines:
                clean_line = line.strip().replace("```", "")
                
                # 1. 헤더 정보 수집
                if clean_line.startswith("HEADER:"):
                    # 중복 방지 (페이지마다 헤더가 나오니까 첫 페이지만 저장하거나 중복 체크)
                    info = clean_line.replace("HEADER:", "").strip()
                    if info and info not in all_header_lines:
                        all_header_lines.append(info)
                
                # 2. 테이블 데이터 수집 (파이프가 있는 줄)
                elif "|" in clean_line:
                    parts = clean_line.split('|')
                    
                    # 칸 개수 강제 맞춤 (10개) - 이게 칸 밀림 방지 핵심!
                    if len(parts) < 10:
                        parts += [""] * (10 - len(parts))
                    elif len(parts) > 10:
                        parts = parts[:10] # 10개 넘으면 자름
                    
                    # 앞뒤 공백 제거
                    parts = [p.strip() for p in parts]
                    
                    # 날짜 형식이 있는 줄만 유효한 데이터로 인정 (노이즈 제거)
                    # (예: 20xx로 시작하거나 숫자로 시작하는 경우)
                    if len(parts[0]) > 0 and (parts[0][0].isdigit()):
                        all_table_rows.append(parts)

        # 데이터프레임 생성
        df = pd.DataFrame(all_table_rows, columns=columns)
        
        # --- 엑셀 저장 및 디자인 (헤더 포함) ---
        
        # 1. 엑셀 파일 생성 (Pandas -> ExcelWriter)
        with pd.ExcelWriter(final_output_xls, engine='openpyxl') as writer:
            # (1) 헤더 정보 먼저 쓰기 (A1셀부터 아래로)
            # 헤더용 임시 데이터프레임
            header_df = pd.DataFrame(all_header_lines)
            header_df.to_excel(writer, index=False, header=False, startrow=0)
            
            # (2) 표 데이터 쓰기 (헤더 정보 아래에, 한 줄 띄우고)
            start_row = len(all_header_lines) + 2
            df.to_excel(writer, index=False, startrow=start_row)
            
        # 2. 디자인 적용 (openpyxl)
        wb = load_workbook(final_output_xls)
        ws = wb.active
        
        # 스타일 정의
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        # 표 헤더 스타일 (진한 회색)
        table_header_fill = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")
        table_header_font = Font(bold=True)

        # 전체 순회하면서 디자인 입히기
        for row in ws.iter_rows():
            for cell in row:
                # 상단 헤더 정보 영역 (데이터 표 시작 전)
                if cell.row < start_row + 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    continue
                
                # 여기서부터는 표 데이터 영역
                cell.border = thin_border
                
                # 표의 제목 줄 (Header)
                if cell.row == start_row + 1:
                    cell.fill = table_header_fill
                    cell.font = table_header_font
                    cell.alignment = center_align
                else:
                    # 데이터 행
                    # 금액 열 (E=5, F=6) -> 콤마가 포함된 텍스트일 수 있음
                    if cell.column in [5, 6]: 
                        cell.alignment = right_align
                        # 혹시 콤마가 빠져있다면 숫자로 변환해서 콤마 찍어주기
                        try:
                            if isinstance(cell.value, str):
                                num = float(cell.value.replace(',', ''))
                                cell.value = num
                                cell.number_format = '#,##0'
                        except:
                            pass # 변환 안되면 그냥 둠 (이미 콤마가 있으므로)
                            
                    # 계좌번호 열 (J=10) -> 텍스트 강제
                    elif cell.column == 10:
                        cell.number_format = '@'
                        cell.value = str(cell.value)
                        cell.alignment = center_align
                    else:
                        cell.alignment = center_align

        # A4 용지 설정
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = False

        # 칸 너비 자동 조절
        for column_cells in ws.columns:
            try:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max(10, min(length + 4, 50))
            except:
                pass

        wb.save(final_output_xls)
        
        with open(final_output_xls, "rb") as f:
            data = f.read()
        
        return data, final_output_xls
            
    except Exception as e:
        return None, f"에러 발생: {str(e)}"
    
    finally:
        if os.path.exists(temp_input): 
            try: os.remove(temp_input)
            except: pass
        if os.path.exists(final_output_xls): 
            try: os.remove(final_output_xls)
            except: pass

# --- 실행 로직 ---
if uploaded_files:
    current_file_ids = "".join([f.name + str(f.size) for f in uploaded_files])
    
    if current_file_ids != st.session_state.last_uploaded_ids:
        st.session_state.processed_files = []
        st.session_state.last_uploaded_ids = current_file_ids
        
        clean_up_trash()
        
        progress_bar = st.progress(0, text="분석 중...")
        total = len(uploaded_files)
        
        for idx, file in enumerate(uploaded_files):
            progress_bar.progress(int((idx / total) * 100), text=f"변환 중... ({idx+1}/{total}) : {file.name}")
            
            excel_data, result_msg = process_pdf_smart(file.getbuffer(), file.name)
            
            if excel_data:
                st.session_state.processed_files.append({
                    "name": result_msg,
                    "data": excel_data
                })
            else:
                st.error(f"❌ '{file.name}' 실패: {result_msg}")
        
        progress_bar.progress(100, text="완료!")

# --- 결과 화면 ---
if st.session_state.processed_files:
    st.success(f"총 {len(st.session_state.processed_files)}개 변환 완료")
    st.markdown("---")
    
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zf:
        for f in st.session_state.processed_files:
            zf.writestr(f['name'], f['data'])
            
    st.download_button(
        label="📦 전체 압축 다운로드 (.ZIP)",
        data=zip_buffer.getvalue(),
        file_name="Converted_Files.zip",
        mime="application/zip",
        use_container_width=True
    )
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    for i, f in enumerate(st.session_state.processed_files):
        with st.container():
            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"**📊 {f['name']}**")
            with col2:
                st.download_button(
                    label="다운로드",
                    data=f['data'],
                    file_name=f['name'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"down_{i}",
                    use_container_width=True
                )